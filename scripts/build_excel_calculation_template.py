"""Build the controlled method-validation calculation workbook."""

from __future__ import annotations

import shutil
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


ROOT = Path(__file__).resolve().parents[1]
OUTPUT = ROOT / "excel" / "分析方法验证原始数据计算模板.xlsx"
ARCHIVE = ROOT / "excel" / "archive" / "XM2026124-01原始计算工作簿.xlsx"
PASSWORD = "report-template"
MAX_IMPURITIES = 10

NAVY = "17324D"
BLUE = "DCEAF7"
YELLOW = "FFF2CC"
GREEN = "E2F0D9"
GREY = "E7E6E6"
RED = "FCE4D6"
WHITE = "FFFFFF"
THIN = Side(style="thin", color="B7C4CE")

INDEX_ROWS: list[tuple[str, str, str, str, str]] = []
RULE_ROWS: list[tuple[str, str, str, str]] = []
TARGET_PATHS = {
    "impurities": "$.impurities[*]",
    "systemSuitability": "$.systemSuitability[*]",
    "specificity": "$.specificity[*]",
    "lodLoq": "$.lod[*] / $.loq[*]（按 result_type 拆分）",
    "linearityPreparation": "$.linearityPreparation[*] / $.intermediateLinearityPreparation[*]（按 run_type 拆分）",
    "linearity": "$.linearity[*] / $.intermediateLinearity[*]（按 run_type 拆分）",
    "repeatability": "$.repeatability[*]",
    "intermediatePrecision": "$.intermediatePrecision[*]",
    "accuracy": "$.blankAmount[*] / $.accuracy[*]（按 record_type 拆分）",
    "solutionStability": "$.solutionStability[*]",
    "robustness": "$.robustnessSpecificity[*] / $.robustnessResult[*]",
}


def style_title(sheet, title: str, end_column: int) -> None:
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_column)
    cell = sheet.cell(1, 1, title)
    cell.fill = PatternFill("solid", fgColor=NAVY)
    cell.font = Font(color=WHITE, bold=True, size=14)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    sheet.row_dimensions[1].height = 30


def style_headers(sheet, row: int, columns: int, english: bool = False) -> None:
    fill = NAVY if not english else BLUE
    color = WHITE if not english else NAVY
    for cell in sheet[row][:columns]:
        cell.fill = PatternFill("solid", fgColor=fill)
        cell.font = Font(color=color, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=THIN)
    sheet.row_dimensions[row].height = 30


def set_widths(sheet, widths: Iterable[int]) -> None:
    for index, width in enumerate(widths, 1):
        sheet.column_dimensions[get_column_letter(index)].width = width


def protect(sheet) -> None:
    sheet.protection.set_password(PASSWORD)
    sheet.protection.sheet = True
    sheet.protection.selectLockedCells = False
    sheet.protection.selectUnlockedCells = True
    sheet.protection.insertRows = False
    sheet.protection.deleteRows = False
    sheet.protection.insertColumns = False
    sheet.protection.deleteColumns = False
    sheet.protection.formatCells = False
    sheet.protection.formatRows = False
    sheet.protection.formatColumns = False


def unlock(cell, input_fill: bool = True) -> None:
    cell.protection = Protection(locked=False)
    if input_fill:
        cell.fill = PatternFill("solid", fgColor=YELLOW)


def formula(cell, value: str, number_format: str = "0.00") -> None:
    cell.value = value
    cell.fill = PatternFill("solid", fgColor=GREEN)
    cell.number_format = number_format
    cell.protection = Protection(locked=True)


def add_numeric_validation(sheet, cells: Iterable, minimum: str = "0", maximum: str = "1E+100") -> None:
    validation = DataValidation(
        type="decimal", operator="between", formula1=minimum, formula2=maximum,
        allow_blank=True, showErrorMessage=True, errorTitle="数据类型错误", error="该字段只能填写规定范围内的数值。",
    )
    sheet.add_data_validation(validation)
    for cell in cells:
        validation.add(cell)


def add_impurity_validation(sheet, cells: Iterable) -> None:
    validation = DataValidation(
        type="list", formula1="='杂质清单'!$A$4:$A$13", allow_blank=False,
        showErrorMessage=True, errorTitle="杂质编号错误", error="请选择杂质清单中的 IMP-01 至 IMP-10。",
    )
    sheet.add_data_validation(validation)
    for cell in cells:
        validation.add(cell)


def add_sequence_validation(sheet, cells: Iterable, maximum: int) -> None:
    validation = DataValidation(
        type="whole", operator="between", formula1="1", formula2=str(maximum), allow_blank=True,
        showErrorMessage=True, errorTitle="序号错误", error=f"序号必须是 1 至 {maximum} 的整数。",
    )
    sheet.add_data_validation(validation)
    for cell in cells:
        validation.add(cell)


def add_table_sheet(workbook: Workbook, title: str, code: str, headers: list[tuple[str, str]],
                    rows: int, widths: list[int], rule: str) -> tuple[object, range]:
    sheet = workbook.create_sheet(title)
    style_title(sheet, title, len(headers))
    for index, (field, label) in enumerate(headers, 1):
        sheet.cell(2, index, field)
        sheet.cell(3, index, label)
    style_headers(sheet, 2, len(headers), True)
    style_headers(sheet, 3, len(headers))
    set_widths(sheet, widths)
    body = range(4, 4 + rows)
    for row in body:
        sheet.row_dimensions[row].height = 22
        for column in range(1, len(headers) + 1):
            sheet.cell(row, column).border = Border(bottom=THIN)
            sheet.cell(row, column).alignment = Alignment(vertical="center", wrap_text=True)
    INDEX_ROWS.append((code, title, f"A2:{get_column_letter(len(headers))}{3 + rows}", "纵向明细表", ",".join(field for field, _ in headers)))
    RULE_ROWS.append((title, f"第4-{3 + rows}行黄色区域", rule, "表头、公式、绿色区域和表外区域均锁定"))
    return sheet, body


def build_guide(workbook: Workbook) -> None:
    sheet = workbook.active
    sheet.title = "模板说明"
    style_title(sheet, "分析方法验证原始数据计算模板", 6)
    metadata = [
        ("模板版本", "2.0"), ("最大杂质数", MAX_IMPURITIES), ("工作簿用途", "验证报告原始数据与计算结果"),
        ("输入颜色", "浅黄色"), ("公式颜色", "浅绿色"), ("保护密码", PASSWORD),
    ]
    for row, (label, value) in enumerate(metadata, 3):
        sheet.cell(row, 1, label).font = Font(bold=True, color=NAVY)
        sheet.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
        sheet.cell(row, 2, value)
    instructions = [
        "先在“杂质清单”登记最多10个杂质，再到各结果表通过 impurity_id 下拉选择。",
        "黄色单元格是唯一允许录入的位置；绿色单元格由公式计算，不能覆盖。",
        "每行只代表一条测定记录；不同杂质、不同序列或不同条件不得写在同一行。",
        "不要新增列、改写英文字段名或在表外记录数据；行数不足时由管理员扩展模板。",
        "本模板不录入结论、验证项目名称、实验过程、接受标准、仪器或溶液配制信息。",
    ]
    sheet.cell(11, 1, "使用顺序")
    sheet.cell(11, 1).font = Font(bold=True, color=WHITE)
    sheet.cell(11, 1).fill = PatternFill("solid", fgColor=NAVY)
    sheet.merge_cells("A11:F11")
    for row, text in enumerate(instructions, 12):
        sheet.cell(row, 1, f"{row - 11}.")
        sheet.cell(row, 2, text)
        sheet.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
        sheet.row_dimensions[row].height = 30
    set_widths(sheet, [16, 24, 18, 18, 18, 18])
    protect(sheet)


def build_impurities(workbook: Workbook) -> None:
    headers = [
        ("impurity_id", "杂质编号"), ("impurity_name", "杂质名称"), ("short_name", "简称"),
        ("limit_ppm", "限度（ppm）"), ("limit_concentration", "限度浓度（ng/ml）"),
        ("api_concentration", "供试品API浓度（mg/ml）"), ("enabled", "启用"), ("note", "备注"),
    ]
    sheet, body = add_table_sheet(workbook, "杂质清单", "impurities", headers, MAX_IMPURITIES,
                                  [14, 28, 16, 16, 23, 25, 12, 28], "固定10个杂质槽位；编号不可修改，其余黄色字段按项目填写")
    yes_no = DataValidation(type="list", formula1='"是,否"', allow_blank=False)
    sheet.add_data_validation(yes_no)
    numeric = []
    for index, row in enumerate(body, 1):
        sheet.cell(row, 1, f"IMP-{index:02d}")
        for column in range(2, 9):
            unlock(sheet.cell(row, column))
        sheet.cell(row, 7, "是" if index <= 3 else "否")
        yes_no.add(sheet.cell(row, 7))
        numeric.extend(sheet.cell(row, column) for column in (4, 5, 6))
    add_numeric_validation(sheet, numeric)
    protect(sheet)


def build_system_suitability(workbook: Workbook) -> None:
    headers = [
        ("impurity_id", "杂质编号"), ("sequence", "No."), ("retention_time", "保留时间（min）"),
        ("peak_area", "峰面积"), ("retention_time_rsd", "保留时间RSD（%）"), ("peak_area_rsd", "峰面积RSD（%）"),
    ]
    sheet, body = add_table_sheet(workbook, "系统适用性结果", "systemSuitability", headers, 60,
                                  [14, 10, 20, 18, 21, 18], "每个启用杂质固定6行，最多10个杂质，共60行")
    numeric = []
    for index, row in enumerate(body):
        start = 4 + (index // 6) * 6
        sheet.cell(row, 1, f"IMP-{index // 6 + 1:02d}")
        sheet.cell(row, 2, index % 6 + 1)
        for column in (3, 4): unlock(sheet.cell(row, column))
        numeric.extend((sheet.cell(row, 3), sheet.cell(row, 4)))
        formula(sheet.cell(row, 5), f'=IFERROR(ROUND(STDEV(C{start}:C{start + 5})/AVERAGE(C{start}:C{start + 5})*100,2),"")')
        formula(sheet.cell(row, 6), f'=IFERROR(ROUND(STDEV(D{start}:D{start + 5})/AVERAGE(D{start}:D{start + 5})*100,2),"")')
    add_numeric_validation(sheet, numeric); protect(sheet)


def build_specificity(workbook: Workbook) -> None:
    headers = [("impurity_id", "杂质编号"), ("solution_name", "溶液名称"), ("retention_time", "保留时间（min）"), ("peak_area", "峰面积")]
    sheet, body = add_table_sheet(workbook, "专属性结果", "specificity", headers, 100, [14, 32, 20, 18], "每个杂质最多10种溶液，每行一个杂质-溶液结果")
    numeric = []
    for index, row in enumerate(body):
        sheet.cell(row, 1, f"IMP-{index // 10 + 1:02d}")
        for column in range(2, 5): unlock(sheet.cell(row, column))
        numeric.extend((sheet.cell(row, 3), sheet.cell(row, 4)))
    add_numeric_validation(sheet, numeric); protect(sheet)


def build_lod_loq(workbook: Workbook) -> None:
    headers = [
        ("impurity_id", "杂质编号"), ("result_type", "结果类型"), ("sequence", "No."), ("sn_ratio", "信噪比S/N"),
        ("peak_area", "峰面积"), ("concentration", "浓度（ng/ml）"), ("sample_content", "相当于供试品中含量（ppm）"),
        ("limit_percentage", "占限度百分比（%）"), ("peak_area_rsd", "峰面积RSD（%）"),
    ]
    sheet, body = add_table_sheet(workbook, "检测限定量限结果", "lodLoq", headers, 90,
                                  [14, 14, 10, 16, 18, 20, 27, 23, 20], "每个杂质3行LOD和6行LOQ；结果类型必须选择LOD或LOQ")
    numeric = []
    for index, row in enumerate(body):
        within = index % 9
        group_start = 4 + (index // 9) * 9
        sheet.cell(row, 1, f"IMP-{index // 9 + 1:02d}")
        sheet.cell(row, 2, "LOD" if within < 3 else "LOQ")
        sheet.cell(row, 3, within + 1 if within < 3 else within - 2)
        for column in range(4, 7): unlock(sheet.cell(row, column))
        numeric.extend(sheet.cell(row, column) for column in (4, 5, 6))
        formula(sheet.cell(row, 7), f'=IFERROR(ROUND(F{row}/VLOOKUP(A{row},\'杂质清单\'!$A$4:$F$13,6,FALSE),2),"")')
        formula(sheet.cell(row, 8), f'=IFERROR(ROUND(G{row}/VLOOKUP(A{row},\'杂质清单\'!$A$4:$F$13,4,FALSE)*100,2),"")')
        formula(sheet.cell(row, 9), f'=IF(B{row}<>"LOQ","",IFERROR(ROUND(STDEV(E{group_start + 3}:E{group_start + 8})/AVERAGE(E{group_start + 3}:E{group_start + 8})*100,2),""))')
    add_numeric_validation(sheet, numeric); protect(sheet)


def build_linearity(workbook: Workbook) -> None:
    headers = [
        ("impurity_id", "杂质编号"), ("run_type", "线性批次"), ("level", "浓度级别"), ("actual_concentration", "实际浓度（ng/ml）"),
        ("peak_area", "峰面积"), ("predicted_peak_area", "预测峰面积"), ("residual", "残差"),
        ("relative_residual", "相对残差（%）"), ("slope", "斜率"), ("intercept", "截距"),
        ("correlation_coefficient", "相关系数r"), ("intercept_ratio", "截距/100%峰面积（%）"),
    ]
    sheet, body = add_table_sheet(workbook, "线性与范围结果", "linearity", headers, 100,
                                  [14, 18, 14, 22, 18, 20, 16, 19, 16, 16, 18, 24], "每个杂质的初始线性和中间精密度线性各固定C1-C5五行")
    numeric = []
    for index, row in enumerate(body):
        block = index // 5; start = 4 + block * 5
        sheet.cell(row, 1, f"IMP-{block // 2 + 1:02d}")
        sheet.cell(row, 2, "INITIAL" if block % 2 == 0 else "INTERMEDIATE")
        sheet.cell(row, 3, f"C{index % 5 + 1}")
        for column in (4, 5): unlock(sheet.cell(row, column))
        numeric.extend((sheet.cell(row, 4), sheet.cell(row, 5)))
        formula(sheet.cell(row, 9), f'=IFERROR(SLOPE(E{start}:E{start + 4},D{start}:D{start + 4}),"")')
        formula(sheet.cell(row, 10), f'=IFERROR(INTERCEPT(E{start}:E{start + 4},D{start}:D{start + 4}),"")')
        formula(sheet.cell(row, 11), f'=IFERROR(CORREL(E{start}:E{start + 4},D{start}:D{start + 4}),"")', "0.0000")
        formula(sheet.cell(row, 6), f'=IFERROR(I{row}*D{row}+J{row},"")', "0")
        formula(sheet.cell(row, 7), f'=IFERROR(E{row}-F{row},"")', "0")
        formula(sheet.cell(row, 8), f'=IFERROR(ROUND(G{row}/F{row}*100,2),"")')
        formula(sheet.cell(row, 12), f'=IFERROR(ROUND(J{row}/E{start + 2}*100,2),"")')
    add_numeric_validation(sheet, numeric); protect(sheet)


def build_linearity_preparation(workbook: Workbook) -> None:
    headers = [
        ("impurity_id", "杂质编号"), ("run_type", "线性批次"), ("solution_name", "溶液名称"),
        ("stock_concentration", "中间贮备液浓度"), ("aliquot_volume", "量取体积（ml）"),
        ("final_volume", "定容至（ml）"), ("theoretical_concentration", "目标溶液理论浓度"),
        ("limit_percentage", "占限度浓度百分比（%）"),
    ]
    sheet, body = add_table_sheet(workbook, "线性溶液数值", "linearityPreparation", headers, 100,
                                  [14, 18, 20, 22, 18, 18, 24, 26], "每个杂质的初始和中间精密度线性各填写C1-C5数值")
    numeric = []
    for index, row in enumerate(body):
        block = index // 5
        sheet.cell(row, 1, f"IMP-{block // 2 + 1:02d}")
        sheet.cell(row, 2, "INITIAL" if block % 2 == 0 else "INTERMEDIATE")
        sheet.cell(row, 3, f"C{index % 5 + 1}")
        for column in range(4, 7): unlock(sheet.cell(row, column))
        numeric.extend(sheet.cell(row, column) for column in (4, 5, 6))
        formula(sheet.cell(row, 7), f'=IFERROR(D{row}*E{row}/F{row},"")')
        formula(sheet.cell(row, 8), f'=IFERROR(G{row}/VLOOKUP(A{row},\'杂质清单\'!$A$4:$F$13,5,FALSE)*100,"")')
    add_numeric_validation(sheet, numeric); protect(sheet)


def build_precision_sheet(workbook: Workbook, title: str, code: str, intermediate: bool) -> None:
    headers = [("impurity_id", "杂质编号")]
    if intermediate: headers.append(("operator_or_day", "人员/日期组"))
    headers += [
        ("sequence", "No."), ("gross_weight", "毛重（mg）"), ("tare_weight", "扣重（mg）"),
        ("api_weight", "API质量（mg）"), ("retention_time", "保留时间（min）"), ("peak_area", "峰面积"),
        ("measured_concentration", "测得浓度（ng/ml）"), ("sample_content", "相当于供试品中含量（ppm）"),
        ("retention_time_rsd", "保留时间RSD（%）"), ("content_rsd", "含量RSD（%）"),
        ("content_ci_low", "含量95%置信下限"), ("content_ci_high", "含量95%置信上限"),
    ]
    rows = 120 if intermediate else 60
    sheet, body = add_table_sheet(workbook, title, code, headers, rows, [14] + ([18] if intermediate else []) + [10, 16, 15, 18, 20, 18, 23, 24, 21, 18, 21, 21],
                                  "每个杂质每组固定6份；中间精密度使用A/B两组，共12份")
    offset = 1 if intermediate else 0; numeric = []
    for index, row in enumerate(body):
        input_end = 9 + offset
        block = index // 6; start = 4 + block * 6
        sheet.cell(row, 1, f"IMP-{block // (2 if intermediate else 1) + 1:02d}")
        if intermediate: sheet.cell(row, 2, "A" if block % 2 == 0 else "B")
        sheet.cell(row, 2 + offset, index % 6 + 1)
        for column in range(3 + offset, input_end + 1): unlock(sheet.cell(row, column))
        numeric.extend(sheet.cell(row, column) for column in range(3 + offset, input_end + 1))
        gross, tare, weight = 3 + offset, 4 + offset, 5 + offset
        rt, concentration = 6 + offset, 8 + offset
        content = 9 + offset; rt_rsd = 10 + offset; content_rsd = 11 + offset
        formula(sheet.cell(row, weight), f'=IFERROR({get_column_letter(gross)}{row}-{get_column_letter(tare)}{row},"")')
        formula(sheet.cell(row, content), f'=IFERROR(ROUND({get_column_letter(concentration)}{row}*20/{get_column_letter(weight)}{row},2),"")')
        formula(sheet.cell(row, rt_rsd), f'=IFERROR(ROUND(STDEV({get_column_letter(rt)}{start}:{get_column_letter(rt)}{start + 5})/AVERAGE({get_column_letter(rt)}{start}:{get_column_letter(rt)}{start + 5})*100,2),"")')
        formula(sheet.cell(row, content_rsd), f'=IFERROR(ROUND(STDEV({get_column_letter(content)}{start}:{get_column_letter(content)}{start + 5})/AVERAGE({get_column_letter(content)}{start}:{get_column_letter(content)}{start + 5})*100,2),"")')
        count = f'COUNT({get_column_letter(content)}{start}:{get_column_letter(content)}{start + 5})'
        avg = f'AVERAGE({get_column_letter(content)}{start}:{get_column_letter(content)}{start + 5})'
        stdev = f'STDEV({get_column_letter(content)}{start}:{get_column_letter(content)}{start + 5})'
        formula(sheet.cell(row, 12 + offset), f'=IFERROR({avg}-TINV(0.05,{count}-1)*{stdev}/SQRT({count}),"")')
        formula(sheet.cell(row, 13 + offset), f'=IFERROR({avg}+TINV(0.05,{count}-1)*{stdev}/SQRT({count}),"")')
    add_numeric_validation(sheet, numeric); protect(sheet)


def build_accuracy(workbook: Workbook) -> None:
    headers = [
        ("impurity_id", "杂质编号"), ("record_type", "记录类型"), ("level", "加标水平"), ("sequence", "No."),
        ("api_weight", "API质量（mg）"), ("peak_area", "峰面积"), ("measured_concentration", "测得浓度（ng/ml）"),
        ("measured_amount", "测得量"), ("blank_amount", "空白量"), ("added_amount", "加入量"),
        ("recovery", "回收率（%）"), ("average_recovery", "平均回收率（%）"), ("recovery_rsd", "回收率RSD（%）"),
    ]
    sheet, body = add_table_sheet(workbook, "准确度结果", "accuracy", headers, 150,
                                  [14, 15, 14, 10, 18, 18, 23, 16, 16, 16, 16, 20, 18], "每个杂质先填空白本底，再按LOQ/50%/100%/150%各3份填写")
    numeric = []
    levels = ("LOQ", "50%", "100%", "150%")
    for index, row in enumerate(body):
        within = index % 15; start = 4 + (index // 15) * 15
        sheet.cell(row, 1, f"IMP-{index // 15 + 1:02d}")
        sheet.cell(row, 2, "BLANK" if within < 3 else "SPIKED")
        sheet.cell(row, 3, "" if within < 3 else levels[(within - 3) // 3])
        sheet.cell(row, 4, within + 1 if within < 3 else (within - 3) % 3 + 1)
        for column in range(5, 8): unlock(sheet.cell(row, column))
        numeric.extend(sheet.cell(row, column) for column in range(5, 8))
        formula(sheet.cell(row, 8), f'=IFERROR(G{row}*20/E{row},"")')
        formula(sheet.cell(row, 9), f'=IF(B{row}="BLANK",H{row},IFERROR(AVERAGE(H{start}:H{start + 2}),""))')
        if within >= 3: unlock(sheet.cell(row, 10)); numeric.append(sheet.cell(row, 10))
        formula(sheet.cell(row, 11), f'=IF(OR(B{row}<>"SPIKED",J{row}=""),"",IFERROR((H{row}-I{row})/J{row}*100,""))')
        formula(sheet.cell(row, 12), f'=IFERROR(AVERAGE(K{start + 3}:K{start + 14}),"")')
        formula(sheet.cell(row, 13), f'=IFERROR(STDEV(K{start + 3}:K{start + 14})/L{row}*100,"")')
    add_numeric_validation(sheet, numeric); protect(sheet)


def build_stability(workbook: Workbook) -> None:
    headers = [
        ("impurity_id", "杂质编号"), ("time_point", "时间点（h）"),
        ("reference_concentration", "对照品溶液浓度"), ("reference_ratio", "对照品与0h比值（%）"),
        ("spiked_concentration", "加标供试品溶液浓度"), ("spiked_ratio", "加标与0h比值（%）"),
    ]
    sheet, body = add_table_sheet(workbook, "溶液稳定性结果", "solutionStability", headers, 100,
                                  [14, 16, 24, 24, 27, 22], "每个杂质最多10个时间点，必须包含0h基准")
    numeric = []
    for index, row in enumerate(body):
        start = 4 + (index // 10) * 10
        sheet.cell(row, 1, f"IMP-{index // 10 + 1:02d}")
        for column in (1, 2, 3, 5): unlock(sheet.cell(row, column))
        sheet.cell(row, 1).protection = Protection(locked=True); sheet.cell(row, 1).fill = PatternFill("solid", fgColor=GREY)
        numeric.extend(sheet.cell(row, column) for column in (2, 3, 5))
        formula(sheet.cell(row, 4), f'=IFERROR(C{row}/C{start}*100,"")')
        formula(sheet.cell(row, 6), f'=IFERROR(E{row}/E{start}*100,"")')
    add_numeric_validation(sheet, numeric); protect(sheet)


def build_robustness(workbook: Workbook) -> None:
    headers = [
        ("impurity_id", "杂质编号"), ("column_no", "色谱柱编号"), ("solution_name", "溶液名称"),
        ("level", "浓度级别"), ("actual_concentration", "实际浓度"), ("peak_area", "峰面积"),
        ("spiked_measured_concentration", "加标溶液测得浓度"), ("slope", "线性斜率"),
        ("intercept", "线性截距"), ("correlation_coefficient", "线性相关系数"), ("column_concentration_ratio", "两柱浓度比值（%）"),
    ]
    sheet, body = add_table_sheet(workbook, "耐用性结果", "robustness", headers, 120,
                                  [14, 16, 28, 14, 18, 18, 25, 16, 16, 18, 23], "每个杂质每根色谱柱固定C1-C5和加标溶液数据")
    numeric = []
    levels = ("C1", "C2", "C3", "C4", "C5", "SPIKED")
    for index, row in enumerate(body):
        block = index // 6; start = 4 + block * 6
        sheet.cell(row, 1, f"IMP-{block // 2 + 1:02d}")
        sheet.cell(row, 2, "COLUMN-1" if block % 2 == 0 else "COLUMN-2")
        sheet.cell(row, 4, levels[index % 6])
        unlock(sheet.cell(row, 3))
        for column in range(5, 8): unlock(sheet.cell(row, column))
        numeric.extend(sheet.cell(row, column) for column in (5, 6, 7))
        formula(sheet.cell(row, 8), f'=IFERROR(SLOPE(F{start}:F{start + 4},E{start}:E{start + 4}),"")')
        formula(sheet.cell(row, 9), f'=IFERROR(INTERCEPT(F{start}:F{start + 4},E{start}:E{start + 4}),"")')
        formula(sheet.cell(row, 10), f'=IFERROR(CORREL(F{start}:F{start + 4},E{start}:E{start + 4}),"")', "0.0000")
        pair_start = 4 + (block // 2) * 12
        formula(sheet.cell(row, 11), f'=IFERROR(G{pair_start + 11}/G{pair_start + 5}*100,"")')
    add_numeric_validation(sheet, numeric); protect(sheet)


def add_index_and_rules(workbook: Workbook) -> None:
    index = workbook.create_sheet("AI数据索引", 2)
    index.append(["table_code", "工作表", "数据区域", "结构类型", "稳定字段", "目标标准JSON路径"])
    for row in INDEX_ROWS: index.append((*row, TARGET_PATHS[row[0]]))
    style_headers(index, 1, 6); set_widths(index, [28, 28, 24, 18, 90, 66])
    for row in index.iter_rows(min_row=2):
        for cell in row: cell.alignment = Alignment(vertical="top", wrap_text=True)
    protect(index)
    rules = workbook.create_sheet("录入规则", 3)
    rules.append(["工作表", "允许录入区域", "填写规则", "禁止操作"])
    for row in RULE_ROWS: rules.append(row)
    style_headers(rules, 1, 4); set_widths(rules, [28, 30, 76, 60])
    for row in rules.iter_rows(min_row=2):
        for cell in row: cell.alignment = Alignment(vertical="top", wrap_text=True)
        rules.row_dimensions[row[0].row].height = 44
    protect(rules)


def add_duplicate_warning(workbook: Workbook) -> None:
    for sheet in workbook.worksheets:
        if sheet.title in {"模板说明", "杂质清单", "AI数据索引", "录入规则"}:
            continue
        end = sheet.max_row
        sheet.conditional_formatting.add(
            f"A4:A{end}", FormulaRule(formula=[f'AND(A4<>"",COUNTIF($A$4:$A${end},A4)>20)'], fill=PatternFill("solid", fgColor=RED)),
        )


def build() -> None:
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    ARCHIVE.parent.mkdir(parents=True, exist_ok=True)
    if OUTPUT.exists() and not ARCHIVE.exists():
        shutil.copy2(OUTPUT, ARCHIVE)
    workbook = Workbook()
    build_guide(workbook)
    build_impurities(workbook)
    build_system_suitability(workbook)
    build_specificity(workbook)
    build_lod_loq(workbook)
    build_linearity_preparation(workbook)
    build_linearity(workbook)
    build_precision_sheet(workbook, "重复性结果", "repeatability", False)
    build_precision_sheet(workbook, "中间精密度结果", "intermediatePrecision", True)
    build_accuracy(workbook)
    build_stability(workbook)
    build_robustness(workbook)
    add_index_and_rules(workbook)
    add_duplicate_warning(workbook)
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.calculation.calcMode = "auto"
    temp = OUTPUT.with_suffix(".tmp.xlsx")
    workbook.save(temp)
    temp.replace(OUTPUT)


def verify() -> None:
    workbook = load_workbook(OUTPUT, data_only=False)
    assert len(workbook.sheetnames) == 14
    assert all(sheet.freeze_panes is None for sheet in workbook.worksheets)
    assert all(sheet.protection.sheet for sheet in workbook.worksheets)
    assert workbook["杂质清单"]["A13"].value == "IMP-10"
    assert workbook["系统适用性结果"]["A4"].protection.locked is True
    assert workbook["系统适用性结果"]["C4"].protection.locked is False
    assert workbook["系统适用性结果"]["E4"].protection.locked is True


if __name__ == "__main__":
    build()
    verify()
