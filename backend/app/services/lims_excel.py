import json
import re
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any

from lxml import html
from openpyxl import load_workbook


EXPECTED_PREFIX = ["ID", "INSTANCEID", "TEMPLATEUNITID", "TYPE", "UNITTITLE", "PARENTID", "UNITBODY"]
STRUCTURED_TYPES = {"Sample", "Standard", "Equipment", "Chromatogram", "Reagent", "Weighing"}


def _identifier(value: Any) -> str | None:
    if value is None or value == "":
        return None
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


def _iso_date(value: Any) -> str | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.isoformat()
    return str(value)


def _parse_body(value: Any) -> dict[str, Any] | None:
    if not isinstance(value, str) or not value.lstrip().startswith("{"):
        return None
    try:
        parsed = json.loads(value)
    except json.JSONDecodeError:
        return None
    return parsed if isinstance(parsed, dict) else None


def _body_items(value: Any) -> list[dict[str, Any]]:
    parsed = _parse_body(value)
    if not parsed:
        return []
    data = parsed.get("data")
    if isinstance(data, list):
        return [item for item in data if isinstance(item, dict)]
    return [data] if isinstance(data, dict) else []


def _plain_text(value: Any) -> str:
    if not isinstance(value, str) or not value.strip():
        return ""
    try:
        root = html.fragment_fromstring(value, create_parent="div")
        return re.sub(r"\s+", " ", root.text_content()).strip()
    except (ValueError, TypeError):
        return re.sub(r"\s+", " ", value).strip()


def _image_urls(value: Any) -> list[str]:
    if not isinstance(value, str) or "<img" not in value.lower():
        return []
    try:
        root = html.fragment_fromstring(value, create_parent="div")
        return [str(url) for url in root.xpath(".//img/@src") if url]
    except (ValueError, TypeError):
        return []


def _evidence(row: dict[str, Any], path: list[str]) -> dict[str, Any]:
    return {
        "type": "LIMS",
        "instanceId": row["instance_id"],
        "unitId": row["id"],
        "unitType": row["type"],
        "sectionPath": path,
        "excelRow": row["excel_row"],
    }


def _normalized_item(unit_type: str, item: dict[str, Any], evidence: dict[str, Any]) -> dict[str, Any]:
    ext = item.get("ext$") if isinstance(item.get("ext$"), dict) else {}
    common = {"sourceRecordId": item.get("lesRecordId") or item.get("id"), "evidence": evidence}
    if unit_type == "Sample":
        return {**common, "sampleName": item.get("sampleName"), "batchNo": item.get("batchNo"),
                "specification": ext.get("spackagetype"), "clientName": item.get("clientName"),
                "remark": item.get("additionalContent"), "sampleNumber": item.get("sampleNumber"),
                "manufacturer": item.get("manufactor"), "appearance": ext.get("samplecolor")}
    if unit_type == "Standard":
        return {**common, "name": ext.get("mtlname") or ext.get("stockmtlname"),
                "content": ext.get("content") or ext.get("purity") or ext.get("titer"),
                "batchNo": item.get("batchNo"), "manufacturer": item.get("manufacturerVendorId"),
                "expiryDate": item.get("validDate"), "stockNo": item.get("stockNo")}
    if unit_type == "Equipment":
        return {**common, "instrumentName": item.get("equiptName"), "model": item.get("specification"),
                "assetNo": item.get("equiptNo"), "manufacturer": item.get("prodVendorId"),
                "calibrationExpiryDate": ext.get("checkvalidate"), "location": item.get("loc")}
    if unit_type == "Chromatogram":
        return {**common, "name": ext.get("model") or ext.get("stationaryphase"),
                "specification": item.get("chromatogramSpec"), "serialNo": item.get("chromatogramNo"),
                "manufacturer": item.get("vendorName"), "stationaryPhase": ext.get("stationaryphase")}
    if unit_type == "Reagent":
        return {**common, "name": ext.get("mtlname") or ext.get("stockmtlname"),
                "grade": ext.get("mtllevel"), "batchNo": item.get("batchNo"),
                "manufacturer": item.get("manufacturerVendorId"), "expiryDate": item.get("validDate")}
    return {**common, "name": ext.get("mtlname") or ext.get("stockmtlname"),
            "batchNo": item.get("batchNo"), "weight": item.get("weight"),
            "weightUnit": item.get("weightUnit"), "weightDate": item.get("weightDate"),
            "equipmentName": item.get("equiptName"), "equipmentNo": item.get("equiptNo"),
            "responseValue": item.get("responseValue")}


def _read_rows(path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        iterator = sheet.iter_rows(values_only=True)
        headers = [str(value or "").strip().upper() for value in next(iterator)]
        if headers[:len(EXPECTED_PREFIX)] != EXPECTED_PREFIX:
            raise ValueError("Excel 列结构不符合 LIMS SQL 结果格式，前七列应为 " + ", ".join(EXPECTED_PREFIX))
        rows = []
        for excel_row, values in enumerate(iterator, start=2):
            row = list(values) + [None] * max(0, 49 - len(values))
            if not any(value not in (None, "") for value in row):
                continue
            rows.append({
                "id": _identifier(row[0]), "instance_id": _identifier(row[1]), "template_unit_id": _identifier(row[2]),
                "type": str(row[3] or ""), "title": str(row[4] or ""), "parent_id": _identifier(row[5]),
                "body": row[6], "order_no": _identifier(row[7]), "excel_row": excel_row,
                "unit_created_by": row[10], "unit_created_time": _iso_date(row[11]),
                "unit_updated_by": row[15], "unit_updated_time": _iso_date(row[16]),
                "record_id": _identifier(row[20]), "template_id": _identifier(row[21]), "record_version": row[22],
                "record_created_by": row[24], "record_created_time": _iso_date(row[25]),
                "record_updated_by": row[29], "record_updated_time": _iso_date(row[30]),
                "process_instance_id": _identifier(row[31]), "last_audited_flag": row[32],
                "last_audited_by": row[34], "last_audited_time": _iso_date(row[35]),
                "submitted_by": row[37], "submitted_time": _iso_date(row[38]),
                "approved_by": row[40], "approved_time": _iso_date(row[41]),
                "project_id": _identifier(row[46]), "source_id": _identifier(row[47]), "source_type": row[48],
            })
        return rows
    finally:
        workbook.close()


def _instance_payload(instance_id: str, rows: list[dict[str, Any]], include_details: bool) -> dict[str, Any]:
    by_id = {row["id"]: row for row in rows if row["id"]}

    def path_for(row: dict[str, Any]) -> list[str]:
        path: list[str] = []
        parent_id = row["parent_id"]
        seen: set[str] = set()
        while parent_id and parent_id not in seen:
            seen.add(parent_id)
            parent = by_id.get(parent_id)
            if not parent:
                break
            if parent["title"]:
                path.insert(0, parent["title"])
            parent_id = parent["parent_id"]
        return path

    first = rows[0]
    title_row = next((row for row in rows if row["type"] == "Section" and row["title"].startswith("实验名称")), None)
    title = title_row["title"].split("：", 1)[-1].strip() if title_row else ""
    groups: dict[str, list[dict[str, Any]]] = defaultdict(list)
    structured_counts: Counter[str] = Counter()
    rich_texts = []
    raw_structured = []
    sections = []
    for row in rows:
        section_path = path_for(row)
        if row["type"] == "Section":
            sections.append({"id": row["id"], "parentId": row["parent_id"], "title": row["title"],
                             "orderNo": row["order_no"], "evidence": _evidence(row, section_path)})
        elif row["type"] in STRUCTURED_TYPES:
            for item in _body_items(row["body"]):
                evidence = _evidence(row, section_path)
                groups[row["type"]].append(_normalized_item(row["type"], item, evidence))
                raw_structured.append({"unitType": row["type"], "data": item, "evidence": evidence})
                structured_counts[row["type"]] += 1
        elif row["type"] == "RichText":
            rich_texts.append({"id": row["id"], "sectionPath": section_path, "plainText": _plain_text(row["body"]),
                               "html": row["body"] if isinstance(row["body"], str) else "",
                               "imageUrls": _image_urls(row["body"]), "evidence": _evidence(row, section_path)})

    approval = []
    for role, name, date in (
        ("编制", first["record_created_by"], first["record_created_time"]),
        ("提交", first["submitted_by"], first["submitted_time"]),
        ("审核", first["last_audited_by"], first["last_audited_time"]),
        ("批准", first["approved_by"], first["approved_time"]),
    ):
        if name:
            approval.append({"field1": role, "field2": "", "field3": name, "field4": "", "date": date})

    summary = {
        "instanceId": instance_id, "projectId": first["project_id"], "title": title,
        "version": first["record_version"], "createdBy": first["record_created_by"],
        "createdTime": first["record_created_time"], "approvedBy": first["approved_by"],
        "approvedTime": first["approved_time"], "rowCount": len(rows),
        "unitCounts": dict(Counter(row["type"] for row in rows)), "structuredDataCounts": dict(structured_counts),
        "richTextCount": len(rich_texts), "richTextCharacters": sum(len(item["plainText"]) for item in rich_texts),
    }
    if not include_details:
        return summary
    return {
        **summary,
        "project": {"id": first["project_id"], "name": title},
        "document": {"code": instance_id, "version": str(first["record_version"] or 0)},
        "approval": approval,
        "samples": groups["Sample"], "referenceStandards": groups["Standard"],
        "instruments": groups["Equipment"], "columns": groups["Chromatogram"],
        "reagents": groups["Reagent"], "weighings": groups["Weighing"],
        "sections": sections, "richTexts": rich_texts, "rawStructured": raw_structured,
    }


def parse_lims_workbook(path: Path, instance_id: str | None = None) -> dict[str, Any]:
    rows = _read_rows(path)
    grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        if row["instance_id"]:
            grouped[row["instance_id"]].append(row)
    if instance_id:
        if instance_id not in grouped:
            raise KeyError(instance_id)
        return _instance_payload(instance_id, grouped[instance_id], True)
    instances = [_instance_payload(key, value, False) for key, value in grouped.items()]
    projects = sorted({item["projectId"] for item in instances if item["projectId"]})
    return {"rowCount": len(rows), "instanceCount": len(instances), "projects": projects, "instances": instances}
