import base64
import subprocess
import tempfile
from pathlib import Path
from urllib.parse import unquote

from lxml import html
from openpyxl import load_workbook


class ExcelChartError(ValueError):
    pass


def _chart_row(chart: object) -> int:
    anchor = getattr(chart, "anchor", None)
    marker = getattr(anchor, "_from", None)
    if marker is None:
        raise ExcelChartError("Excel 图表缺少可识别的锚点")
    return int(marker.row)


def _image_row(image: object) -> int | None:
    table, node = None, image.getparent()
    while node is not None:
        if node.tag == "table":
            table = node
            break
        node = node.getparent()
    if table is None:
        return None
    tr, node = None, image.getparent()
    while node is not None:
        if node.tag == "tr":
            tr = node
            break
        node = node.getparent()
    if tr is None:
        return None
    return list(table).index(tr)


def _rendered_chart_images(path: Path) -> list[tuple[int, bytes]]:
    """返回每张图表图片的 (所在表格行号, 图片字节)，按 HTML 文档顺序。"""
    with tempfile.TemporaryDirectory(prefix="excel-charts-") as directory:
        output_dir = Path(directory)
        profile = output_dir / "profile"
        result = subprocess.run(
            ["libreoffice", "--headless", f"-env:UserInstallation={profile.as_uri()}",
             "--convert-to", "html", "--outdir", str(output_dir), str(path)],
            capture_output=True, text=True, timeout=90, check=False,
        )
        html_files = list(output_dir.glob("*.html"))
        if result.returncode or len(html_files) != 1:
            message = (result.stderr or result.stdout or "未生成 HTML").strip()
            raise ExcelChartError(f"LibreOffice 图表渲染失败：{message}")
        document = html.fromstring(html_files[0].read_bytes())
        images: list[tuple[int, bytes]] = []
        for image in document.xpath("//img"):
            reference = image.get("src")
            image_path = output_dir / Path(unquote(str(reference))).name
            if not (image_path.exists() and image_path.suffix.lower() == ".png"):
                continue
            row = _image_row(image)
            if row is None:
                raise ExcelChartError("无法定位图表图片的行位置，不能可靠地配对图片与图表")
            images.append((row, image_path.read_bytes()))
        return images


def extract_residual_chart_values(path: Path, points_per_test: int = 5) -> list[str]:
    workbook = load_workbook(path, data_only=False, read_only=False, keep_vba=True)
    if "线性" not in workbook.sheetnames:
        raise ExcelChartError("Excel 缺少“线性”工作表")
    charts = sorted(workbook["线性"]._charts, key=_chart_row)
    images = _rendered_chart_images(path)
    if len(images) != len(charts):
        raise ExcelChartError(f"Excel 图表对象共 {len(charts)} 个，但仅渲染出 {len(images)} 张图片")
    # “隔一张取一张”默认渲染顺序与图表行序一致；若不一致会静默配错图（报告里图表张冠李戴），
    # 因此行号必须非递减，否则直接失败而不是猜
    rows = [row for row, _ in images]
    if any(current < previous for previous, current in zip(rows, rows[1:])):
        raise ExcelChartError("图表图片渲染顺序与图表行序不一致，无法可靠配对图片")
    normal_images = [image for index, (_, image) in enumerate(images) if index % 2 == 0]
    values: list[str] = []
    for image in normal_images:
        data_url = "data:image/png;base64," + base64.b64encode(image).decode("ascii")
        values.extend([data_url] * points_per_test)
    return values
