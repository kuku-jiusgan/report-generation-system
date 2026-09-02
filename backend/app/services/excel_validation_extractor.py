import hashlib
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell.cell import TYPE_ERROR

from .excel_rule_engine import excel_display_value


RESULT_SHEETS = {
    "首页", "系统适用性", "对照品配置", "专属性", "检测限与定量限", "线性",
    "重复性跟中间精密度", "准确度", "溶液稳定性", "耐用性", "检测",
}


def _value(value: Any) -> Any:
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return value


class ValidationWorkbookReader:
    def __init__(self, path: Path) -> None:
        self.path = path
        self.formulas = load_workbook(path, data_only=False, read_only=False, keep_vba=True)
        self.values = load_workbook(path, data_only=True, read_only=False, keep_vba=True)
        self.warnings: list[str] = []

    def cell(self, sheet: str, address: str, required: bool = False) -> Any:
        formula_cell = self.formulas[sheet][address]
        value_cell = self.values[sheet][address]
        value = excel_display_value(_value(value_cell.value), formula_cell.number_format)
        missing_formula = formula_cell.data_type == "f" and value in (None, "")
        invalid = value_cell.data_type == TYPE_ERROR or (isinstance(value, str) and value.startswith("#"))
        if missing_formula or invalid:
            self.warnings.append(f"{sheet}!{address} 的公式缓存无有效结果")
            return None
        if required and value in (None, ""):
            self.warnings.append(f"{sheet}!{address} 未填写")
        return value

    def evidence(self, sheet: str, address: str) -> dict[str, Any]:
        return {"sheet": sheet, "cell": address}

    def validate(self) -> None:
        missing = sorted(RESULT_SHEETS.difference(self.values.sheetnames))
        if missing:
            raise ValueError(f"缺少验证结果计算页：{', '.join(missing)}")

    def impurity_names(self) -> list[str]:
        raw = self.cell("首页", "B8", True)
        try:
            count = int(raw)
        except (TypeError, ValueError) as error:
            raise ValueError("首页!B8 必须是 1 至 15 的杂质数量") from error
        if not 1 <= count <= 15:
            raise ValueError("首页!B8 必须是 1 至 15 的杂质数量")
        names = [str(self.cell("首页", f"B{9 + index}", True) or "").strip() for index in range(count)]
        if any(not name for name in names):
            raise ValueError("首页杂质名称不能为空")
        return names

    def extract(self) -> dict[str, Any]:
        self.validate()
        names = self.impurity_names()
        payload: dict[str, Any] = {
            "project": {"name": self.cell("首页", "B3", True)},
            "document": {"version": self.cell("首页", "F4")},
            "impurity": [{"impurityName": name} for name in names],
            "referenceStandards": self._reference_standards(),
            "systemSuitability": self._system_suitability(names),
            "specificity": self._specificity(names),
            "limit": self._limits(names),
            "robustnessSpecificity": self._robustness(),
            "conclusions": self._conclusions(names),
        }
        payload["validationResults"] = self._validation_results(names)
        payload["_meta"] = {
            "format": "WENXIA_VALIDATION_V49", "impurityCount": len(names),
            "impurityNames": names, "warnings": self.warnings,
            "sha256": hashlib.sha256(self.path.read_bytes()).hexdigest(),
        }
        return payload

    def _reference_standards(self) -> list[dict[str, Any]]:
        result = []
        ws = self.values["对照品配置"]
        for row in range(3, ws.max_row + 1):
            name = self.cell("对照品配置", f"A{row}")
            if name not in (None, ""):
                result.append({"name": name, "content": self.cell("对照品配置", f"C{row}"),
                               "_evidence": self.evidence("对照品配置", f"A{row}:C{row}")})
        return result

    def _system_suitability(self, names: list[str]) -> list[dict[str, Any]]:
        result = []
        for index, name in enumerate(names):
            start_col = 1 + index * 3
            for sequence in range(1, 7):
                row = 2 + sequence
                result.append({
                    "impurityName": name, "solutionName": self.cell("系统适用性", f"{_col(start_col)}{row}"),
                    "sequence": sequence,
                    "retentionTime": self.cell("系统适用性", f"{_col(start_col + 1)}{row}"),
                    "peakArea": self.cell("系统适用性", f"{_col(start_col + 2)}{row}"),
                    "_evidence": self.evidence("系统适用性", f"{_col(start_col)}{row}:{_col(start_col + 2)}{row}"),
                })
        return result

    def _specificity(self, names: list[str]) -> list[dict[str, Any]]:
        result = []
        for index, name in enumerate(names):
            top = 1 + index * 5
            for offset in range(1, 5):
                row = top + offset
                result.append({
                    "impurityName": name, "solutionName": self.cell("专属性", f"B{row}"),
                    "retentionTime": self.cell("专属性", f"C{row}"),
                    "peakArea": self.cell("专属性", f"D{row}"),
                    "_evidence": self.evidence("专属性", f"A{row}:D{row}"),
                })
        return result

    def _limits(self, names: list[str]) -> list[dict[str, Any]]:
        return [{
            "impurityName": name, "field4": self.cell("检测限与定量限", f"A{3 + index}"),
            "field5": self.cell("检测限与定量限", f"B{3 + index}"),
            "_evidence": self.evidence("检测限与定量限", f"A{3 + index}:I{3 + index}"),
        } for index, name in enumerate(names)]

    def _robustness(self) -> list[dict[str, Any]]:
        result = []
        for row in range(2, 8):
            name = self.cell("耐用性", f"A{row}")
            if name not in (None, ""):
                result.append({"solutionName": name, "field2": self.cell("耐用性", f"B{row}"),
                               "field3": self.cell("耐用性", f"C{row}"),
                               "_evidence": self.evidence("耐用性", f"A{row}:C{row}")})
        return result

    def _conclusions(self, names: list[str]) -> list[dict[str, Any]]:
        locations = [("系统适用性", 10, 13, "horizontal"), ("专属性", 5, 5, "vertical"),
                     ("线性", 11, 24, "vertical"), ("准确度", 28, 28, "vertical")]
        result = []
        for sheet, offset, block, direction in locations:
            for index, name in enumerate(names):
                address = f"{_col(2 + index * 3)}{offset}" if direction == "horizontal" else f"E{offset + index * block}"
                value = self.cell(sheet, address)
                if value not in (None, ""):
                    result.append({"text": str(value), "validationItem": sheet, "impurityName": name,
                                   "_evidence": self.evidence(sheet, address)})
        return result

    def _validation_results(self, names: list[str]) -> dict[str, Any]:
        return {"impurityNames": names, "sheets": {
            name: _sheet_matrix(self.values[name], self.formulas[name]) for name in RESULT_SHEETS if name != "首页"
        }}


def _col(index: int) -> str:
    result = ""
    while index:
        index, remainder = divmod(index - 1, 26)
        result = chr(65 + remainder) + result
    return result


def _sheet_matrix(value_sheet: Any, formula_sheet: Any) -> list[list[Any]]:
    rows = []
    for row in value_sheet.iter_rows():
        values = []
        for cell in row:
            formula_cell = formula_sheet[cell.coordinate]
            value = None if cell.data_type == TYPE_ERROR else _value(cell.value)
            values.append(excel_display_value(value, formula_cell.number_format))
        if any(value not in (None, "") for value in values):
            rows.append(values)
    return rows


def extract_validation_workbook(path: Path) -> dict[str, Any]:
    return ValidationWorkbookReader(path).extract()
