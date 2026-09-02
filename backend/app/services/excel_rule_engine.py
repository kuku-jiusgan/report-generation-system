import hashlib
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell.cell import TYPE_ERROR


class ExcelRuleError(ValueError):
    pass


def _column(index: int) -> str:
    result = ""
    while index:
        index, remainder = divmod(index - 1, 26)
        result = chr(65 + remainder) + result
    return result


def _value(value: Any) -> Any:
    return value.isoformat() if isinstance(value, (date, datetime)) else value


def _decimal_places(number_format: str) -> int | None:
    if not number_format or number_format == "General":
        return None
    section = number_format.split(";", 1)[0]
    if "." not in section:
        return 0 if "0" in section or "#" in section else None
    decimal_part = section.split(".", 1)[1]
    count = 0
    for char in decimal_part:
        if char in {"0", "#"}:
            count += 1
            continue
        break
    return count


def excel_display_value(value: Any, number_format: str) -> Any:
    if not isinstance(value, float):
        return value
    places = _decimal_places(number_format)
    if places is None:
        return value
    rounded = round(value, places)
    return int(rounded) if places == 0 else rounded


class WorkbookValues:
    def __init__(self, path: Path) -> None:
        self.formulas = load_workbook(path, data_only=False, read_only=False, keep_vba=True)
        self.values = load_workbook(path, data_only=True, read_only=False, keep_vba=True)
        self.warnings: list[str] = []

    def read(self, sheet: str, row: int, column: int, required: bool = False) -> Any:
        if sheet not in self.values.sheetnames:
            if required:
                raise ExcelRuleError(f"缺少必填工作表：{sheet}")
            self.warnings.append(f"缺少可选工作表：{sheet}")
            return None
        formula, cached = self.formulas[sheet].cell(row, column), self.values[sheet].cell(row, column)
        value = excel_display_value(_value(cached.value), formula.number_format)
        invalid = cached.data_type == TYPE_ERROR or (isinstance(value, str) and value.startswith("#"))
        if invalid or (formula.data_type == "f" and value in (None, "")):
            message = f"{sheet}!{_column(column)}{row} 的公式缓存无有效结果"
            if required:
                raise ExcelRuleError(message)
            self.warnings.append(message)
            return None
        if required and value in (None, ""):
            raise ExcelRuleError(f"{sheet}!{_column(column)}{row} 未填写")
        return value


def _fixed(reader: WorkbookValues, rule: dict[str, Any]) -> Any:
    return reader.read(str(rule["sheet"]), int(rule["row"]), int(rule["column"]), bool(rule.get("required")))


def _repeat(reader: WorkbookValues, rule: dict[str, Any]) -> list[dict[str, Any]]:
    sheet = str(rule["sheet"])
    count_rule = rule.get("count") or {}
    count = int(_fixed(reader, count_rule)) if count_rule else int(rule.get("repeatCount", 0))
    if count < 0 or count > int(rule.get("maxRepeat", 100)):
        raise ExcelRuleError(f"集合 {rule['collection']} 的循环次数 {count} 超出限制")
    rows = rule.get("rows") or {}
    row_start, row_end = int(rows.get("start", 1)), int(rows.get("end", 1))
    if row_end < row_start or row_end - row_start > 1000:
        raise ExcelRuleError(f"集合 {rule['collection']} 的数据行范围无效")
    start_column, column_step = int(rule.get("startColumn", 1)), int(rule.get("columnStep", 0))
    start_row, row_step = int(rule.get("startRow", 1)), int(rule.get("rowStep", 0))
    records: list[dict[str, Any]] = []
    for repeat_index in range(count):
        for row_index, source_row in enumerate(range(row_start, row_end + 1)):
            record: dict[str, Any] = {}
            evidence: list[str] = []
            for field in rule.get("fields", []):
                mode = str(field.get("mode", "CELL"))
                if mode == "INDEX":
                    value = row_index + int(field.get("base", 1))
                elif mode == "REPEAT_VALUE":
                    source = field.get("source") or {}
                    value = reader.read(str(source["sheet"]), int(source["row"]) + repeat_index * int(source.get("rowStep", 0)),
                                        int(source["column"]) + repeat_index * int(source.get("columnStep", 0)), bool(field.get("required")))
                else:
                    row = source_row + start_row - row_start + repeat_index * row_step + int(field.get("rowOffset", 0))
                    column = start_column + repeat_index * column_step + int(field.get("columnOffset", 0))
                    value = reader.read(sheet, row, column, bool(field.get("required")))
                    evidence.append(f"{sheet}!{_column(column)}{row}")
                record[str(field["name"])] = value
            record["_evidence"] = {"sheet": sheet, "cells": evidence, "repeatIndex": repeat_index}
            if any(value not in (None, "", record["_evidence"]) for value in record.values()):
                records.append(record)
    return records


def _place(payload: dict[str, Any], output: str, value: Any) -> None:
    # 点号路径一律按层级落位，FIXED 与 REPEAT_BLOCK 保持一致；
    # 否则 REPEAT_BLOCK 配了 "a.b" 会产出字面量为 "a.b" 的扁平键，和 FIXED 的嵌套结构对不上
    current = payload
    parts = output.split(".")
    for part in parts[:-1]:
        current = current.setdefault(part, {})
    current[parts[-1]] = value


def execute_excel_rules(path: Path, snapshot: dict[str, Any], version_id: int | None = None) -> dict[str, Any]:
    reader = WorkbookValues(path)
    payload: dict[str, Any] = {}
    for rule in snapshot.get("rules", []):
        if not rule.get("enabled", True):
            continue
        output = str(rule["output"])
        if rule["kind"] == "FIXED":
            _place(payload, output, _fixed(reader, rule))
        elif rule["kind"] == "REPEAT_BLOCK":
            _place(payload, output, _repeat(reader, rule))
        else:
            raise ExcelRuleError(f"不支持的 Excel 规则类型：{rule['kind']}")
    payload["_meta"] = {"format": snapshot.get("code", "CUSTOM"), "ruleVersionId": version_id,
                        "warnings": reader.warnings, "sha256": hashlib.sha256(path.read_bytes()).hexdigest()}
    return payload


def validate_excel_snapshot(snapshot: dict[str, Any]) -> list[str]:
    errors: list[str] = []
    if not snapshot.get("rules"):
        errors.append("至少需要一条 Excel 集合规则")
    outputs: set[str] = set()
    for index, rule in enumerate(snapshot.get("rules", []), 1):
        output = str(rule.get("output") or "")
        if not output:
            errors.append(f"第 {index} 条规则缺少输出路径")
        if output in outputs:
            errors.append(f"输出路径重复：{output}")
        outputs.add(output)
        if rule.get("kind") not in {"FIXED", "REPEAT_BLOCK"}:
            errors.append(f"第 {index} 条规则类型无效")
        if rule.get("kind") == "REPEAT_BLOCK" and not rule.get("fields"):
            errors.append(f"集合 {output} 至少需要一个输出字段")
    return errors
