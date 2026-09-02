import tempfile
from pathlib import Path

import pytest
from openpyxl import Workbook

from backend.app.admin_routes.excel_rules import _collection_sizes
from backend.app.services import excel_chart_extractor
from backend.app.services.excel_rule_engine import execute_excel_rules


WORKBOOK = Path(__file__).parents[2] / "excel" / "文霞-标准计算表-动态-V49-解锁版.xlsm"


def _repeat_block_rule(output: str) -> dict:
    return {
        "kind": "REPEAT_BLOCK", "name": output, "output": output, "sheet": "数据",
        "repeatCount": 2, "maxRepeat": 10, "startColumn": 1, "columnStep": 0,
        "startRow": 2, "rowStep": 1, "rows": {"start": 2, "end": 2},
        "fields": [
            {"name": "a", "mode": "CELL", "columnOffset": 0},
            {"name": "b", "mode": "CELL", "columnOffset": 1},
        ],
        "enabled": True,
    }


def _save_workbook(directory: Path) -> Path:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "数据"
    sheet.cell(2, 1, 10)
    sheet.cell(2, 2, "x")
    sheet.cell(3, 1, 20)
    sheet.cell(3, 2, "y")
    path = directory / "repeat.xlsx"
    workbook.save(path)
    return path


def test_repeat_block_dotted_output_nests_like_fixed() -> None:
    with tempfile.TemporaryDirectory() as directory:
        path = _save_workbook(Path(directory))
        snapshot = {"code": "T", "rules": [_repeat_block_rule("nested.collection")]}
        payload = execute_excel_rules(path, snapshot)

    records = payload["nested"]["collection"]
    assert len(records) == 2
    assert [record["a"] for record in records] == [10, 20]
    assert [record["b"] for record in records] == ["x", "y"]


def test_repeat_block_plain_output_stays_top_level() -> None:
    with tempfile.TemporaryDirectory() as directory:
        path = _save_workbook(Path(directory))
        snapshot = {"code": "T", "rules": [_repeat_block_rule("collection")]}
        payload = execute_excel_rules(path, snapshot)

    assert isinstance(payload["collection"], list)
    assert len(payload["collection"]) == 2


def test_collection_sizes_reports_nested_paths() -> None:
    sizes = _collection_sizes({"a": [1, 2, 3], "b": {"c": [4, 5], "d": "scalar"}, "_meta": {"k": 1}})
    assert sizes == {"a": 3, "b.c": 2, "b.d": 1}


def test_chart_images_out_of_row_order_fail_fast(monkeypatch) -> None:
    # 真实工作簿“线性”表有 4 个图表；渲染行号倒序 → 隔一张取一张会配错图，必须直接失败
    monkeypatch.setattr(
        excel_chart_extractor, "_rendered_chart_images",
        lambda path: [(40, b"p4"), (28, b"p3"), (16, b"p2"), (4, b"p1")],
    )
    with pytest.raises(excel_chart_extractor.ExcelChartError):
        excel_chart_extractor.extract_residual_chart_values(WORKBOOK, 5)


def test_chart_images_in_row_order_pair_by_position(monkeypatch) -> None:
    monkeypatch.setattr(
        excel_chart_extractor, "_rendered_chart_images",
        lambda path: [(4, b"p1"), (16, b"p2"), (28, b"p3"), (40, b"p4")],
    )
    values = excel_chart_extractor.extract_residual_chart_values(WORKBOOK, 2)

    # 4 图表取偶数位（0、2）→ 两张图，各重复 points_per_test=2 次
    assert len(values) == 4
    assert values[0].startswith("data:image/png;base64,")
    assert values[0] == values[1]
    assert values[2] == values[3]
    assert values[0] != values[2]
