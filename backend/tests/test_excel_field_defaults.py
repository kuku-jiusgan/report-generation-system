from pathlib import Path
import tempfile

from openpyxl import Workbook

from backend.app.services.excel_field_extractor import extract_excel_fields
from backend.app.services.excel_rule_defaults import (
    EXCEL_FIELD_PATHS,
    REPEATABILITY_DETAIL_COLUMNS,
    REPEATABILITY_SUMMARY_CELLS,
    _rule_config,
)


WORKBOOK = Path(__file__).parents[2] / "excel" / "文霞-标准计算表-动态-V49-解锁版.xlsm"
FIELD_CODES = [f"uncategorized.field_{index:03d}" for index in range(7, 14)]
QUANTITATION_CODES = [f"uncategorized.field_{index:03d}" for index in range(14, 21)]
LINEARITY_CODES = [f"uncategorized.field_{index:03d}" for index in range(21, 29)]
REPEATABILITY_CODES = [f"uncategorized.field_{index:03d}" for index in range(30, 43)]


def test_extracts_detection_limit_result_columns() -> None:
    fields = [
        {"fieldCode": code, "cardinality": "MANY", "legacyJsonPath": EXCEL_FIELD_PATHS[code]}
        for code in FIELD_CODES
    ]
    rules = [
        {"id": index, "fieldCode": code, "sourceType": "EXCEL", "priority": 50,
         "enabled": True, "config": _rule_config(code, EXCEL_FIELD_PATHS[code])}
        for index, code in enumerate(FIELD_CODES, 1)
    ]

    payload = extract_excel_fields(WORKBOOK, fields, rules)

    assert "field_007" not in payload.get("custom", {})
    assert payload["lod"] == [
        {"name": "杂质D", "field2": 34.02, "field3": 27.43, "field4": 30.65,
         "field5": 1.28, "field6": 1.3, "field7": 5},
        {"name": "杂质A2", "field2": 34.02, "field3": 27.43, "field4": 30.65,
         "field5": None, "field6": None, "field7": None},
    ]


def test_detection_limit_rules_read_one_row_per_impurity() -> None:
    for code, column in zip(FIELD_CODES, range(3, 10), strict=True):
        config = _rule_config(code, EXCEL_FIELD_PATHS[code])
        assert config["sheet"] == "检测限与定量限"
        assert config["rowStart"] == config["rowEnd"] == 3
        assert config["startColumn"] == column
        assert config["rowStep"] == 1
        assert config["repeatCountSource"] == {"sheet": "首页", "row": 8, "column": 2}


def test_extracts_quantitation_limit_result_columns() -> None:
    fields = [
        {"fieldCode": code, "cardinality": "MANY", "legacyJsonPath": EXCEL_FIELD_PATHS[code]}
        for code in QUANTITATION_CODES
    ]
    rules = [
        {"id": index, "fieldCode": code, "sourceType": "EXCEL", "priority": 50,
         "enabled": True, "config": _rule_config(code, EXCEL_FIELD_PATHS[code])}
        for index, code in enumerate(QUANTITATION_CODES, 1)
    ]

    payload = extract_excel_fields(WORKBOOK, fields, rules)

    assert "field_014" not in payload.get("custom", {})
    assert payload["loq"][:2] == [
        {"sequence": 1, "field2": 58.44, "peakArea": 15766, "field4": 2.1,
         "field5": None, "field6": None, "field7": None},
        {"sequence": 2, "field2": 60.69, "peakArea": 16006, "field4": None,
         "field5": None, "field6": None, "field7": None},
    ]
    assert len(payload["loq"]) == 12


def test_quantitation_limit_rules_skip_headers_and_conclusions_between_groups() -> None:
    for code, column in zip(QUANTITATION_CODES, range(4, 11), strict=True):
        config = _rule_config(code, EXCEL_FIELD_PATHS[code])
        assert config["rowStart"] == 8
        assert config["rowEnd"] == 13
        assert config["rowStartOffsetFromRepeatCount"] == 6
        assert config["rowCount"] == 6
        assert config["startColumn"] == column
        assert config["rowStep"] == 8


def test_quantitation_limit_rules_follow_dynamic_single_impurity_layout() -> None:
    workbook = Workbook()
    cover = workbook.active
    cover.title = "首页"
    cover["B8"] = 1
    result = workbook.create_sheet("检测限与定量限")
    rows = [
        [1, 58.44, 15766, 2.0583, 2.57, 2.6, 10.4],
        [2, 60.69, 16006, None, None, None, None],
        [3, 70.11, 15409, None, None, None, None],
        [4, 50.12, 15196, None, None, None, None],
        [5, 59.39, 15968, None, None, None, None],
        [6, 50.72, 15816, None, None, None, None],
    ]
    for row_index, values in enumerate(rows, 7):
        for column_index, value in enumerate(values, 4):
            result.cell(row_index, column_index, value)
    with tempfile.TemporaryDirectory() as directory:
        path = Path(directory) / "single-impurity.xlsx"
        workbook.save(path)
        fields = [{"fieldCode": code, "cardinality": "MANY", "legacyJsonPath": EXCEL_FIELD_PATHS[code]}
                  for code in QUANTITATION_CODES]
        rules = [{"id": index, "fieldCode": code, "sourceType": "EXCEL", "priority": 50,
                  "enabled": True, "config": _rule_config(code, EXCEL_FIELD_PATHS[code])}
                 for index, code in enumerate(QUANTITATION_CODES, 1)]

        payload = extract_excel_fields(path, fields, rules)

    assert payload["loq"][0] == {
        "sequence": 1, "field2": 58.44, "peakArea": 15766, "field4": 2.0583,
        "field5": 2.57, "field6": 2.6, "field7": 10.4,
    }
    assert len(payload["loq"]) == 6


def test_extracts_horizontal_linearity_results_and_statistics() -> None:
    fields = [
        {"fieldCode": code, "cardinality": "MANY", "legacyJsonPath": EXCEL_FIELD_PATHS[code]}
        for code in LINEARITY_CODES
    ]
    rules = [
        {"id": index, "fieldCode": code, "sourceType": "EXCEL", "priority": 50,
         "enabled": True, "config": _rule_config(code, EXCEL_FIELD_PATHS[code])}
        for index, code in enumerate(LINEARITY_CODES, 1)
    ]

    payload = extract_excel_fields(WORKBOOK, fields, rules)

    assert "field_021" not in payload.get("custom", {})
    assert payload["linearity"][:2] == [
        {"solutionName": "C1", "field2": 2.57, "peakArea": 14889,
         "regressionEquation": "y = 5886.7751x + 969.9591", "correlationCoefficient": 0.999918,
         "interceptRatio": 0.64, "predictedPeakArea": 14944, "residual": -55},
        {"solutionName": "C2", "field2": 12.84, "peakArea": 78112,
         "regressionEquation": "y = 5886.7751x + 969.9591", "correlationCoefficient": 0.999918,
         "interceptRatio": 0.64, "predictedPeakArea": 76208, "residual": 1904},
    ]
    assert len(payload["linearity"]) == 10


def test_extracts_residual_charts_as_embedded_png_values() -> None:
    code = "uncategorized.field_029"
    fields = [{"fieldCode": code, "cardinality": "MANY", "legacyJsonPath": EXCEL_FIELD_PATHS[code]}]
    rules = [{"id": 1, "fieldCode": code, "sourceType": "EXCEL", "priority": 50,
              "enabled": True, "config": _rule_config(code, EXCEL_FIELD_PATHS[code])}]

    payload = extract_excel_fields(WORKBOOK, fields, rules)

    assert len(payload["linearity"]) == 10
    assert payload["linearity"][0]["residualChart"].startswith("data:image/png;base64,")
    assert payload["linearity"][0]["residualChart"] == payload["linearity"][4]["residualChart"]
    assert payload["linearity"][0]["residualChart"] != payload["linearity"][5]["residualChart"]


def test_extracts_repeatability_detail_and_summary_fields() -> None:
    fields = [{"fieldCode": code, "cardinality": "MANY", "legacyJsonPath": EXCEL_FIELD_PATHS[code]}
              for code in REPEATABILITY_CODES]
    rules = [{"id": index, "fieldCode": code, "sourceType": "EXCEL", "priority": 50,
              "enabled": True, "config": _rule_config(code, EXCEL_FIELD_PATHS[code])}
             for index, code in enumerate(REPEATABILITY_CODES, 1)]
    payload = extract_excel_fields(WORKBOOK, fields, rules)

    assert payload["custom"]["field_030"][:6] == [1, 2, 3, 4, 5, 6]
    assert payload["custom"]["field_031"][:6] == [20.02, 20.34, 20.17, 20.13, 20.16, 20.31]
    assert payload["custom"]["field_032"][:6] == [8.211, 8.212, 8.208, 8.207, 8.209, 8.21]
    assert payload["custom"]["field_033"][:6] == [156531, 160279, 154371, 159010, 161875, 160824]
    assert payload["custom"]["field_034"][:6] == [26.3, 26.93, 25.94, 26.72, 27.2, 27.02]
    assert payload["custom"]["field_035"][:6] == [26.3, 26.5, 25.7, 26.5, 27, 26.6]
    assert payload["custom"]["field_036"] == [0.1, 0.1]
    assert payload["custom"]["field_037"] == [25.985, None]
    assert payload["custom"]["field_038"] == [26.882, None]
    assert payload["custom"]["field_039"] == [1.06115360384027, None]
    assert payload["custom"]["field_040"] == [4.1694437805321, None]
    assert payload["custom"]["field_041"] == [103.939, None]
    assert payload["custom"]["field_042"] == [107.527, None]


def test_repeatability_rules_use_six_row_blocks_and_summary_cells() -> None:
    for code, column in REPEATABILITY_DETAIL_COLUMNS.items():
        config = _rule_config(code, EXCEL_FIELD_PATHS[code])
        assert config["rowStart"] == 3
        assert config["rowEnd"] == 8
        assert config["startColumn"] == column
        assert config["rowStep"] == 33

    weighing = _rule_config(
        "uncategorized.field_031", EXCEL_FIELD_PATHS["uncategorized.field_031"]
    )
    assert weighing["workbookLocation"]["valueColumn"] == "E（称样量）"

    for code, (row, column) in REPEATABILITY_SUMMARY_CELLS.items():
        config = _rule_config(code, EXCEL_FIELD_PATHS[code])
        assert config["rowStart"] == config["rowEnd"] == row
        assert config["startColumn"] == column
        assert config["rowStep"] == 33
